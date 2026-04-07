"""Unit tests for src.datalab.pipeline.ingest."""

import csv
import json
import tempfile
from pathlib import Path

import pytest

from src.datalab.pipeline.ingest import (
    FileInfo,
    detect_encoding,
    identify_format,
)


# ── Fixtures ─────────────────────────────────────────────


@pytest.fixture
def utf8_csv(tmp_path) -> Path:
    """Create a UTF-8 CSV file with header + data rows."""
    p = tmp_path / "data.csv"
    p.write_text(
        "name,age,city\nAlice,30,Seoul\nBob,25,Busan\nCarol,28,Daegu\n",
        encoding="utf-8",
    )
    return p


@pytest.fixture
def euckr_csv(tmp_path) -> Path:
    """Create a EUC-KR (cp949) encoded CSV file."""
    p = tmp_path / "korean.csv"
    p.write_text(
        "이름,나이,도시\n홍길동,30,서울\n김철수,25,부산\n",
        encoding="cp949",
    )
    return p


@pytest.fixture
def tsv_file(tmp_path) -> Path:
    """Create a TSV file."""
    p = tmp_path / "data.tsv"
    p.write_text(
        "name\tage\tcity\nAlice\t30\tSeoul\nBob\t25\tBusan\n",
        encoding="utf-8",
    )
    return p


@pytest.fixture
def xlsx_file(tmp_path) -> Path:
    """Create a simple Excel file with openpyxl."""
    import openpyxl

    p = tmp_path / "data.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["name", "age", "city"])
    ws.append(["Alice", 30, "Seoul"])
    ws.append(["Bob", 25, "Busan"])
    ws.append(["Carol", 28, "Daegu"])
    # Add a second sheet
    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["id", "value"])
    ws2.append([1, 100])
    wb.save(p)
    return p


@pytest.fixture
def xlsx_with_merged(tmp_path) -> Path:
    """Create an Excel file with merged cells."""
    import openpyxl

    p = tmp_path / "merged.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Header A", "Header B", "Header C"])
    ws.append([1, 2, 3])
    ws.append([4, 5, 6])
    ws.merge_cells("A1:B1")
    ws.merge_cells("A2:A3")
    wb.save(p)
    return p


@pytest.fixture
def json_file(tmp_path) -> Path:
    """Create a JSON array-of-objects file."""
    p = tmp_path / "data.json"
    data = [
        {"name": "Alice", "age": 30, "city": "Seoul"},
        {"name": "Bob", "age": 25, "city": "Busan"},
    ]
    p.write_text(json.dumps(data), encoding="utf-8")
    return p


@pytest.fixture
def pdf_file(tmp_path) -> Path:
    """Create a dummy PDF file (just the extension matters)."""
    p = tmp_path / "report.pdf"
    p.write_bytes(b"%PDF-1.4 dummy")
    return p


@pytest.fixture
def image_file(tmp_path) -> Path:
    """Create a dummy PNG file."""
    p = tmp_path / "chart.png"
    p.write_bytes(b"\x89PNG\r\n\x1a\n dummy image")
    return p


# ── Tests: detect_encoding ───────────────────────────────


class TestDetectEncoding:
    def test_utf8(self, utf8_csv):
        enc = detect_encoding(utf8_csv)
        assert enc.lower().replace("-", "") in ("utf8", "ascii", "utf-8", "utf8")

    def test_euckr_normalized_to_cp949(self, euckr_csv):
        enc = detect_encoding(euckr_csv)
        assert enc == "cp949"

    def test_nonexistent_file_returns_utf8(self, tmp_path):
        """Non-existent file should fall back to utf-8."""
        fake = tmp_path / "nope.csv"
        enc = detect_encoding(fake)
        assert enc == "utf-8"


# ── Tests: identify_format — CSV ─────────────────────────


class TestIdentifyCSV:
    def test_format(self, utf8_csv):
        info = identify_format(utf8_csv)
        assert info.format == "csv"

    def test_column_names(self, utf8_csv):
        info = identify_format(utf8_csv)
        assert info.column_names == ["name", "age", "city"]

    def test_columns_count(self, utf8_csv):
        info = identify_format(utf8_csv)
        assert info.columns == 3

    def test_rows_count(self, utf8_csv):
        info = identify_format(utf8_csv)
        assert info.rows == 3  # 3 data rows (header excluded)

    def test_sample_rows(self, utf8_csv):
        info = identify_format(utf8_csv)
        assert len(info.sample_rows) <= 5
        assert info.sample_rows[0] == ["Alice", "30", "Seoul"]

    def test_encoding_field(self, utf8_csv):
        info = identify_format(utf8_csv)
        assert info.encoding.lower().replace("-", "") in ("utf8", "ascii", "utf-8")


# ── Tests: identify_format — TSV ─────────────────────────


class TestIdentifyTSV:
    def test_format(self, tsv_file):
        info = identify_format(tsv_file)
        assert info.format == "tsv"

    def test_column_names(self, tsv_file):
        info = identify_format(tsv_file)
        assert info.column_names == ["name", "age", "city"]


# ── Tests: identify_format — Excel ───────────────────────


class TestIdentifyExcel:
    def test_format(self, xlsx_file):
        info = identify_format(xlsx_file)
        assert info.format == "excel"

    def test_sheet_names(self, xlsx_file):
        info = identify_format(xlsx_file)
        assert info.sheet_names == ["Sheet1", "Sheet2"]
        assert info.sheets == 2

    def test_columns(self, xlsx_file):
        info = identify_format(xlsx_file)
        assert info.column_names == ["name", "age", "city"]
        assert info.columns == 3

    def test_rows(self, xlsx_file):
        info = identify_format(xlsx_file)
        assert info.rows == 3  # 3 data rows on first sheet

    def test_sample_rows(self, xlsx_file):
        info = identify_format(xlsx_file)
        assert len(info.sample_rows) >= 1
        assert info.sample_rows[0] == ["Alice", "30", "Seoul"]

    def test_merged_cells(self, xlsx_with_merged):
        info = identify_format(xlsx_with_merged)
        assert info.merged_cells_count == 2


# ── Tests: identify_format — JSON ────────────────────────


class TestIdentifyJSON:
    def test_format(self, json_file):
        info = identify_format(json_file)
        assert info.format == "json"

    def test_columns_from_keys(self, json_file):
        info = identify_format(json_file)
        assert set(info.column_names) == {"name", "age", "city"}
        assert info.columns == 3

    def test_rows_count(self, json_file):
        info = identify_format(json_file)
        assert info.rows == 2

    def test_json_object_not_array(self, tmp_path):
        """A JSON object (not array) should still parse."""
        p = tmp_path / "obj.json"
        p.write_text(json.dumps({"key": "value"}), encoding="utf-8")
        info = identify_format(p)
        assert info.format == "json"
        assert info.rows == 0  # not an array of objects


# ── Tests: identify_format — PDF / Image / Text ──────────


class TestIdentifyOther:
    def test_pdf(self, pdf_file):
        info = identify_format(pdf_file)
        assert info.format == "pdf"
        assert info.path == pdf_file

    def test_image_png(self, image_file):
        info = identify_format(image_file)
        assert info.format == "image"

    def test_image_jpg(self, tmp_path):
        p = tmp_path / "photo.jpg"
        p.write_bytes(b"\xff\xd8\xff dummy jpeg")
        info = identify_format(p)
        assert info.format == "image"

    def test_image_jpeg(self, tmp_path):
        p = tmp_path / "photo.jpeg"
        p.write_bytes(b"\xff\xd8\xff dummy jpeg")
        info = identify_format(p)
        assert info.format == "image"

    def test_unknown_extension(self, tmp_path):
        p = tmp_path / "data.xyz"
        p.write_text("some content", encoding="utf-8")
        info = identify_format(p)
        assert info.format == "text"
