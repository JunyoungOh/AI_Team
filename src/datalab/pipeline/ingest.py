"""DataLab ingest pipeline — encoding detection + format identification.

Parses uploaded files to extract metadata (column names, row counts,
sample rows, sheet info) that JARVIS uses to understand data structure
before analysis.
"""

from __future__ import annotations

import csv
import io
import json
import logging
from dataclasses import dataclass, field
from pathlib import Path

import chardet

logger = logging.getLogger(__name__)

# ── Korean encoding aliases to normalise → cp949 ────────

_KOREAN_ENCODINGS = frozenset({"euc-kr", "euckr", "cp949", "johab", "iso-2022-kr"})

# ── Format extension maps ────────────────────────────────

_CSV_EXTS = {".csv"}
_TSV_EXTS = {".tsv"}
_EXCEL_EXTS = {".xlsx", ".xls"}
_JSON_EXTS = {".json"}
_PDF_EXTS = {".pdf"}
_IMAGE_EXTS = {".png", ".jpg", ".jpeg"}


# ── FileInfo dataclass ──────────────────────────────────


@dataclass
class FileInfo:
    """Metadata extracted from an ingested file."""

    path: Path
    format: str  # csv, excel, json, tsv, text, pdf, image
    encoding: str = "utf-8"
    rows: int = 0
    columns: int = 0
    column_names: list[str] = field(default_factory=list)
    sheets: int = 0
    sheet_names: list[str] = field(default_factory=list)
    sample_rows: list[list[str]] = field(default_factory=list)
    merged_cells_count: int = 0


# ── Public API ──────────────────────────────────────────


def detect_encoding(path: Path, sample_size: int = 65536) -> str:
    """Detect file encoding using chardet.

    Korean encodings (euc-kr, cp949, johab) are normalised to "cp949".
    Falls back to "utf-8" on error or low confidence.
    """
    try:
        raw = path.read_bytes()[:sample_size]
    except (OSError, FileNotFoundError):
        return "utf-8"

    if not raw:
        return "utf-8"

    result = chardet.detect(raw)
    encoding = (result.get("encoding") or "utf-8").lower()
    confidence = result.get("confidence", 0.0) or 0.0

    # Normalise Korean encodings
    if encoding.replace("-", "") in {e.replace("-", "") for e in _KOREAN_ENCODINGS}:
        return "cp949"

    if confidence < 0.5:
        return "utf-8"

    return encoding


def identify_format(path: Path) -> FileInfo:
    """Identify file format by extension and parse metadata."""
    ext = path.suffix.lower()

    if ext in _CSV_EXTS:
        return _parse_csv(path, fmt="csv")
    if ext in _TSV_EXTS:
        return _parse_csv(path, fmt="tsv")
    if ext in _EXCEL_EXTS:
        return _parse_excel(path)
    if ext in _JSON_EXTS:
        return _parse_json(path)
    if ext in _PDF_EXTS:
        return FileInfo(path=path, format="pdf")
    if ext in _IMAGE_EXTS:
        return FileInfo(path=path, format="image")

    return FileInfo(path=path, format="text")


# ── Private parsers ─────────────────────────────────────


def _parse_csv(path: Path, fmt: str = "csv") -> FileInfo:
    """Parse CSV/TSV: detect delimiter, read header + up to 5 sample rows."""
    encoding = detect_encoding(path)

    try:
        raw = path.read_bytes()
        text = raw.decode(encoding)
    except (UnicodeDecodeError, OSError):
        text = raw.decode("utf-8", errors="replace")
        encoding = "utf-8"

    # Detect delimiter
    delimiter = "\t" if fmt == "tsv" else ","
    try:
        sample = text[:8192]
        dialect = csv.Sniffer().sniff(sample)
        delimiter = dialect.delimiter
    except csv.Error:
        pass  # fallback to default

    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows_data: list[list[str]] = []
    for row in reader:
        rows_data.append(row)

    if not rows_data:
        return FileInfo(path=path, format=fmt, encoding=encoding)

    header = rows_data[0]
    data_rows = rows_data[1:]
    sample_rows = data_rows[:5]

    return FileInfo(
        path=path,
        format=fmt,
        encoding=encoding,
        rows=len(data_rows),
        columns=len(header),
        column_names=header,
        sample_rows=sample_rows,
    )


def _parse_excel(path: Path) -> FileInfo:
    """Parse Excel file using openpyxl read_only mode."""
    import openpyxl

    try:
        wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    except Exception as exc:
        logger.warning("Failed to open Excel file %s: %s", path, exc)
        return FileInfo(path=path, format="excel")

    try:
        sheet_names = wb.sheetnames
        sheets = len(sheet_names)

        # Parse first sheet for column/row info
        ws = wb[sheet_names[0]]

        # Count merged cells across all sheets
        merged_count = 0
        for sn in sheet_names:
            sheet = wb[sn]
            merged_count += len(sheet.merged_cells.ranges)

        # Read all rows from first sheet
        all_rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            all_rows.append([str(v) if v is not None else "" for v in row])

        if not all_rows:
            return FileInfo(
                path=path,
                format="excel",
                sheets=sheets,
                sheet_names=sheet_names,
                merged_cells_count=merged_count,
            )

        header = all_rows[0]
        data_rows = all_rows[1:]
        sample_rows = data_rows[:5]

        return FileInfo(
            path=path,
            format="excel",
            encoding="utf-8",
            rows=len(data_rows),
            columns=len(header),
            column_names=header,
            sheets=sheets,
            sheet_names=sheet_names,
            sample_rows=sample_rows,
            merged_cells_count=merged_count,
        )
    finally:
        wb.close()


def _parse_json(path: Path) -> FileInfo:
    """Parse JSON file — detect array-of-objects structure."""
    encoding = detect_encoding(path)

    try:
        raw = path.read_bytes()
        text = raw.decode(encoding)
        data = json.loads(text)
    except (json.JSONDecodeError, UnicodeDecodeError, OSError) as exc:
        logger.warning("Failed to parse JSON file %s: %s", path, exc)
        return FileInfo(path=path, format="json", encoding=encoding)

    # Array of objects → tabular structure
    if isinstance(data, list) and data and isinstance(data[0], dict):
        keys = list(data[0].keys())
        sample_rows = [
            [str(row.get(k, "")) for k in keys] for row in data[:5]
        ]
        return FileInfo(
            path=path,
            format="json",
            encoding=encoding,
            rows=len(data),
            columns=len(keys),
            column_names=keys,
            sample_rows=sample_rows,
        )

    # Non-tabular JSON (object, scalar, etc.)
    return FileInfo(path=path, format="json", encoding=encoding)
